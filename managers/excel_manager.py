from returns.result import Result, Success, Failure
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def read_from_sheet(path_xlsx: str, sheet_name: str, header_row_number: int, *row_names)-> Result[dict, str]:
    ''' Return dictionary with keys as row names and values as list of table values'''
    try:
        wb = load_workbook(path_xlsx)
    except Exception as e:
        return Failure(f"Failed to open Excel file. {str(e)}")
    
    if not sheet_name in wb.sheetnames:
        return Failure(f"Failed to find sheet '{sheet_name}' in Excel file. Existing sheets: {wb.sheetnames}")

    ws = wb[sheet_name]
    header = ws[header_row_number]

    # Get column letters as in Excel
    cols = [None for x in range(len(row_names))]
    for i, s in enumerate(row_names):
        cols[i] = find_column_containing_str(header, s)

    # Get column values as list in dictionary
    col_dict_values = {}
    for row_name, col in zip(row_names, cols):
        cells = ws[col][header_row_number:]
        col_dict_values[row_name] = [entry.value.strip() for entry in cells]

    if len(set(map(len, col_dict_values.values())))!=1:
        return Failure(f"Failure. Columns in Excel document are of different length")

    return Success(col_dict_values)
    

def find_column_containing_str(row, s: str):
    for cell in row:
        if cell.value is not None: 
            if s.lower() in cell.value.lower():
                return get_column_letter(cell.column)
    return None

def find_row_containing_str(col, s: str):
    for cell in col:
        if cell.value is not None: #We need to check that the cell is not empty.
            if s.lower() in cell.value.lower(): #Check if the value of the cell contains the string
                return cell.row
    return None

def find_column_containing_str_excluding_str(row, s_cont: str, s_excl: str):
    for cell in row:
        if cell.value is not None:
            if s_cont.lower() in cell.value.lower() and s_excl.lower() not in cell.value.lower():
                return get_column_letter(cell.column)
    return None